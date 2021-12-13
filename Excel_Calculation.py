import openpyxl

from datetime import *
wb = openpyxl.load_workbook("C:\\Users\\vivek\\Desktop\\MR. MAHESH BASUDKAR PROJECTION REPORT.xlsx")

sheets = wb.sheetnames
s1 = wb['Need Base Analysis']

#s2= wb["Projection Report"]

s2 =wb["Sheet1"]

dictonaryOfData ={}

def calculateCompoundValue(amount, rate, time):
    compoundAmount = amount * (pow((1 + rate / 100), time))
    compoundAmount = round(compoundAmount, 0)
    return int(compoundAmount)

def calculateAge(retirementAge):
    FatherAge = s1['D6'].value
    #vivneedBaseDataDictionary['monthOfDob'] = 6
    currentYear = datetime.now().year
    FatherAge = FatherAge.year
    currentAgeInYear =currentYear - FatherAge
    yearDifference =retirementAge-currentAgeInYear
    print("this is year remaining in difference")
    print(yearDifference)
    return int(yearDifference)

def calculateExpenses():
    global dictonaryOfData
    Gross_Income = s1['E83'].value
    Expense_with_investment = 0
    Investments = 0
    for i in range(86, 97):
        # print(s1['E'+ str(i)].value)
        if s1['E' + str(i)].value is not None:
            if i == 93 or i == 95:
                print("this is i value", i)
                Investments = Investments + int(s1['E' + str(i)].value)
            Expense_with_investment = Expense_with_investment + int(s1['E' + str(i)].value)
    TotalExpense_without_investment = Expense_with_investment - Investments
    Investible_Surplus = Gross_Income - TotalExpense_without_investment
    dictonaryOfData['investibleSurplus'] = Investible_Surplus
    dictonaryOfData['totalExpenseWithoutInvestment']=TotalExpense_without_investment
    dictonaryOfData['retirementYear'] =s1['D46'].value
    if dictonaryOfData['retirementYear'] is not None:
        dictonaryOfData['retirementcost'] =calculateCompoundValue(
            dictonaryOfData['totalExpenseWithoutInvestment'], dictonaryOfData["Rate"], dictonaryOfData['retirementYear'])
    if dictonaryOfData['retirementcost'] is not None:
        new_row = 46
        new_column = 6
        retirementcost = dictonaryOfData["retirementcost"]
        s1.cell(new_row, new_column, value=retirementcost)
        wb.save("C:\\Users\\vivek\\Desktop\\MR. MAHESH BASUDKAR PROJECTION REPORT.xlsx")

def needBaseSheetValue(Rate):
    #return Amount
    global dictonaryOfData
    currentYear = datetime.now().year
    dictonaryOfData["Rate"]=Rate
    numberOfChild=0
    firstChildDob = s1['I14'].value
    secondChildDob = s1['I15'].value
    thirdChildDob = s1['I16'].value
    if(firstChildDob is not None):
        numberOfChild = numberOfChild +1
        firstChildDob = firstChildDob.year
        age =18-(currentYear - firstChildDob)
        dictonaryOfData["firstChildPresentCostForGraduation"]= s1['E22'].value
        dictonaryOfData["firstChildGraduationCost"] =calculateCompoundValue(
            dictonaryOfData["firstChildPresentCostForGraduation"],Rate,age)
        if dictonaryOfData["firstChildGraduationCost"] is not None:
            new_row =22
            new_column =8
            firstChildGraduationCost = dictonaryOfData["firstChildGraduationCost"]
            s1.cell(new_row, new_column, value=firstChildGraduationCost)
            wb.save("C:\\Users\\vivek\\Desktop\\MR. MAHESH BASUDKAR PROJECTION REPORT.xlsx")
        dictonaryOfData["firstChildPresentCostForPostGraduation"] = s1['J22'].value
        dictonaryOfData["firstChildPostGraduationCost"] = calculateCompoundValue(
            dictonaryOfData["firstChildPresentCostForPostGraduation"], Rate, age+4)
        if dictonaryOfData["firstChildPostGraduationCost"] is not None:
            new_row =22
            new_column =11
            firstChildPostGraduationCost = dictonaryOfData["firstChildPostGraduationCost"]
            s1.cell(new_row, new_column, value=firstChildPostGraduationCost)
            wb.save("C:\\Users\\vivek\\Desktop\\MR. MAHESH BASUDKAR PROJECTION REPORT.xlsx")
        dictonaryOfData["firstChildPresentWeddingCost"] = s1['I27'].value
        dictonaryOfData["firstChildWeddingCost"]=calculateCompoundValue(
            dictonaryOfData["firstChildPresentWeddingCost"], Rate, age+7)
        if dictonaryOfData["firstChildWeddingCost"] is not None:
            new_row =27
            new_column =10
            firstChildWeddingCost = dictonaryOfData["firstChildPresentWeddingCost"]
            s1.cell(new_row, new_column, value=firstChildWeddingCost)
            wb.save("C:\\Users\\vivek\\Desktop\\MR. MAHESH BASUDKAR PROJECTION REPORT.xlsx")


    if (secondChildDob is not None):
        numberOfChild = numberOfChild + 1
        secondChildDob = secondChildDob.year
        age = 18 - (currentYear - secondChildDob)
        dictonaryOfData["secondChildPresentCostForGraduation"] = s1['E23'].value
        dictonaryOfData["secondChildGraduationCost"] = calculateCompoundValue(
            dictonaryOfData["secondChildPresentCostForGraduation"], Rate, age)
        if dictonaryOfData["secondChildGraduationCost"] is not None:
            new_row =23
            new_column =8
            secondChildGraduationCost = dictonaryOfData["secondChildGraduationCost"]
            s1.cell(new_row, new_column, value=secondChildGraduationCost)
            wb.save("C:\\Users\\vivek\\Desktop\\MR. MAHESH BASUDKAR PROJECTION REPORT.xlsx")

        dictonaryOfData["secondChildPresentCostForPostGraduation"] = s1['J23'].value
        dictonaryOfData["secondChildPostGraduationCost"] = calculateCompoundValue(
            dictonaryOfData["secondChildPresentCostForPostGraduation"], Rate, age + 4)
        if dictonaryOfData["secondChildPostGraduationCost"] is not None:
            new_row =23
            new_column =11
            secondChildPostGraduationCost = dictonaryOfData["secondChildPostGraduationCost"]
            s1.cell(new_row, new_column, value=secondChildPostGraduationCost)
            wb.save("C:\\Users\\vivek\\Desktop\\MR. MAHESH BASUDKAR PROJECTION REPORT.xlsx")

        dictonaryOfData["secondChildPresentWeddingCost"] = s1['I28'].value
        dictonaryOfData["secondChildWeddingCost"] = calculateCompoundValue(
            dictonaryOfData["secondChildPresentWeddingCost"], Rate, age + 7)
        if dictonaryOfData["secondChildWeddingCost"] is not None:
            new_row =28
            new_column =10
            secondChildWeddingCost = dictonaryOfData["secondChildWeddingCost"]
            s1.cell(new_row, new_column, value=secondChildWeddingCost)
            wb.save("C:\\Users\\vivek\\Desktop\\MR. MAHESH BASUDKAR PROJECTION REPORT.xlsx")

    if (thirdChildDob is not None):
        numberOfChild = numberOfChild + 1
        thirdChildDob=thirdChildDob.year
        age = 18 - (currentYear - secondChildDob)
        dictonaryOfData["thirdChildPresentCostForGraduation"] = s1['E24'].value
        dictonaryOfData["thirdChildGraduationCost"] = calculateCompoundValue(
            dictonaryOfData["thirdChildPresentCostForGraduation"], Rate, age)
        if dictonaryOfData["thirdChildGraduationCost"] is not None:
            new_row =24
            new_column =8
            thirdChildGraduationCost = dictonaryOfData["thirdChildGraduationCost"]
            s1.cell(new_row, new_column, value=thirdChildGraduationCost)
            wb.save("C:\\Users\\vivek\\Desktop\\MR. MAHESH BASUDKAR PROJECTION REPORT.xlsx")

        dictonaryOfData["thirdChildPresentCostForPostGraduation"] = s1['J24'].value
        dictonaryOfData["thirdChildPostGraduationCost"] = calculateCompoundValue(
            dictonaryOfData["thirdChildPresentCostForPostGraduation"], Rate, age + 4)
        if dictonaryOfData["thirdChildPostGraduationCost"] is not None:
            new_row =24
            new_column =11
            thirdChildPostGraduationCost = dictonaryOfData["thirdChildPostGraduationCost"]
            s1.cell(new_row, new_column, value=thirdChildPostGraduationCost)
            wb.save("C:\\Users\\vivek\\Desktop\\MR. MAHESH BASUDKAR PROJECTION REPORT.xlsx")
        dictonaryOfData["thirdChildPresentWeddingCost"] = s1['I29'].value
        dictonaryOfData["thirdChildWeddingCost"] = calculateCompoundValue(
            dictonaryOfData["thirdChildPresentWeddingCost"], Rate, age + 7)
        if dictonaryOfData["thirdChildWeddingCost"] is not None:
            new_row = 29
            new_column = 10
            thirdChildWeddingCost = dictonaryOfData["thirdChildWeddingCost"]
            s1.cell(new_row, new_column, value=thirdChildWeddingCost)
            wb.save("C:\\Users\\vivek\\Desktop\\MR. MAHESH BASUDKAR PROJECTION REPORT.xlsx")

    calculateExpenses()
    return True

