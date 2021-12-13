import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from datetime import *
wb = openpyxl.load_workbook("C:\\Users\\vivek\\Desktop\\MR. MAHESH BASUDKAR PROJECTION REPORT.xlsx")

sheets = wb.sheetnames
s1 = wb['Need Base Analysis']

#s2= wb["Projection Report"]

s2 =wb["Sheet2"]
redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')
totalMonthlyInstallment=0
currentAgeInYear=0
yearsToCalculateInsurance ={}
Rate =0
valueOfInvestmentList = []
currentYear = datetime.now().year
needBaseDataDictionary = {
    "retirementAge" : s1["D46"].value,
    "monthlyRequirementAfterRetirement" : s1["F46"].value,
    "currentYear": currentYear
 }
insuranceCostDataDictionary ={}
s2["D3"].fill =redFill

insuranceExpiryYear ={}
insuranceExpiryAmount ={}


def calculateCompoundValue(amount, rate, time):
    compoundAmount = amount * (pow((1 + rate / 100), time))
    compoundAmount = round(compoundAmount, 0)
    return int(compoundAmount)

def calculateAge():
    global needBaseDataDictionary
    FatherAge = s1['D6'].value

    needBaseDataDictionary['monthOfDob'] = FatherAge.month
    #vivneedBaseDataDictionary['monthOfDob'] = 6
    currentYear = datetime.now().year
    FatherAge = FatherAge.year
    currentAgeInYear =currentYear - FatherAge
    return currentAgeInYear


def getDobOfAllMemebers():
    global needBaseDataDictionary
    global currentYear
    numberOfChild=0
    needBaseDataDictionary['currentAgeInYear']= calculateAge()
    firstChildDob = s1['I14'].value
    secondChildDob = s1['I15'].value
    thirdChildDob = s1['I16'].value
    if(firstChildDob is not None):
        numberOfChild = numberOfChild +1
        firstChildDob =firstChildDob.year
        needBaseDataDictionary['firstChildDob'] = firstChildDob
        print(18-(currentYear-firstChildDob))
        #Here You can change inflation rate at line 61
        needBaseDataDictionary['firstChildGraducationCost'] = calculateCompoundValue(int(s1['E22'].value),10,18-(currentYear-firstChildDob))
        needBaseDataDictionary['firstChildPostGraduationCost'] =calculateCompoundValue(int(s1['J22'].value),10,22-(currentYear-firstChildDob))
        needBaseDataDictionary['firstChildMarriage'] = calculateCompoundValue(int(s1['I27'].value),10,25-(currentYear-firstChildDob))
    if (secondChildDob is not None):
        numberOfChild = numberOfChild + 1
        secondChildDob =secondChildDob.year
        needBaseDataDictionary['secondChildDob'] = secondChildDob
        needBaseDataDictionary['secondChildGraducationCost'] = calculateCompoundValue(int(s1['E23'].value), 10,18 - (currentYear - secondChildDob))
        needBaseDataDictionary['secondChildPostGraduationCost'] = calculateCompoundValue(int(s1['J23'].value), 10, 22 - (currentYear - secondChildDob))
        needBaseDataDictionary['secondChildMarriage'] = calculateCompoundValue(int(s1['I28'].value), 10,25 - (currentYear - secondChildDob))
    if (thirdChildDob is not None):
        numberOfChild = numberOfChild + 1
        thirdChildDob=thirdChildDob.year
        needBaseDataDictionary['thirdChildDob'] = thirdChildDob
        needBaseDataDictionary['thirdChildGraducationCost'] = calculateCompoundValue(int(s1['E24'].value), 10, 18 - (currentYear - thirdChildDob))
        needBaseDataDictionary['thirdChildPostGraduationCost'] = calculateCompoundValue(int(s1['J24'].value), 10,22 - (currentYear - thirdChildDob))
        needBaseDataDictionary['thirdChildMarriage'] = calculateCompoundValue(int(s1['I29'].value), 10,25 - (currentYear - thirdChildDob))


    needBaseDataDictionary['numberOfChild'] =numberOfChild

def calculateExpenses():
    global needBaseDataDictionary
    Gross_Income = s1['E83'].value
    Expense_with_investment = 0
    Investments = 0
    for i in range(86, 97):
        # print(s1['E'+ str(i)].value)
        if s1['E' + str(i)].value is not None:
            if i == 93 or i == 95:
                print("this is i value", i)
                Investments = Investments + int(s1['E' + str(i)].value)
            Expense_with_investment = Expense_with_investment + int(s1['E' + str(i)].value)
    TotalExpense_without_investment = Expense_with_investment - Investments
    Investible_Surplus = Gross_Income - TotalExpense_without_investment
    needBaseDataDictionary['Investible_Surplus'] = Investible_Surplus

def calculateInsuaranceCost():
    global needBaseDataDictionary
    global insuranceCostDataDictionary
    global insuranceExpiryAmount
    listd =[]
    print("insurance cost")
    print(insuranceCostDataDictionary.values())
    for i in range(57, 67):
        # print(s1['E'+ str(i)].value)

        if s1['S' + str(i)].value is not None:
            print("insurance cost222")
            print(insuranceCostDataDictionary.items())

            for key,value in insuranceCostDataDictionary.items():
                if(key==s1['S' + str(i)].value):
                    listd.append(s1['P' + str(i)].value)
                    key = {s1['S' + str(i)].value:listd}
                    insuranceCostDataDictionary.update(key)
            insuranceCostDataDictionary[s1['S' + str(i)].value] = s1['P' + str(i)].value
            #print(insuranceCostDataDictionary)


def insuranceExpiryAmountToAdd():
    global insuranceExpiryAmount
    global  insuranceExpiryYear
    for i in range(57,67):
        if s1['S' + str(i)].value and s1['P' + str(i)].value is not None:
            insuranceExpiryAmount[i] =s1['P' + str(i)].value
            insuranceExpiryYear[i]=s1['S' + str(i)].value
    print("#--------------------")
    print(insuranceExpiryAmount)
    print(insuranceExpiryYear)
    print("#--------------------")



#---------Commenting the following methods--------
#print(getDobOfAllMemebers())
#print(calculateExpenses())
#print(calculateInsuaranceCost())
#print(needBaseDataDictionary)
#print(insuranceCostDataDictionary)
#print(insuranceExpiryAmountToAdd())

# For First column values required till retirement:

def writeFirstColumnTillRetirement(retirementAge,currentYear,currentAgeInYear):
    yearsToRetirement = retirementAge - currentAgeInYear
    global yearsToCalculateInsurance
    j = 4
    k = 1
    count= 1
    for i in range(currentAgeInYear, 86):
        new_row = j
        new_col = k
        s2.cell(new_row, new_col, value=currentYear)
        yearsToCalculateInsurance[currentYear] = new_row

        currentYear = currentYear + 1
        j = j + 1
    wb.save("C:\\Users\\vivek\\Desktop\\MR. MAHESH BASUDKAR PROJECTION REPORT1.xlsx")

def writeSecondColumnTillRetirement(retirementAge,currentAgeInYear):
    j = 4
    k = 2
    count=1
    for i in range(currentAgeInYear, 86):
        new_row = j
        new_col = k
        s2.cell(new_row, new_col, value=i)
        j = j + 1
    wb.save("C:\\Users\\vivek\\Desktop\\MR. MAHESH BASUDKAR PROJECTION REPORT1.xlsx")


def writeThirdColumnTillRetirement(monthlyInvestment, retirementAge, currentAgeInYear):
    monthlyInvestment =monthlyInvestment * 12
    j = 5
    k = 3
    for i in range(currentAgeInYear, retirementAge):
        new_row = j
        new_col = k
        s2.cell(new_row, new_col, value=monthlyInvestment)
        j = j + 1
    for i in range(retirementAge+1, 86):

        new_row = j
        new_col = k
        s2.cell(new_row, new_col, value=0)
        j = j + 1

    wb.save("C:\\Users\\vivek\\Desktop\\MR. MAHESH BASUDKAR PROJECTION REPORT1.xlsx")
def checkAmount(expiryYear,i, insuranceExpiryAmount, currentYear):
    finalYear =expiryYear - currentYear
    j=4+finalYear
    print("dgfgfffg#######")
    print(finalYear)
    print(currentYear)
    print(expiryYear)
    if s2.cell(j,3).value !=0:
        val = s2.cell(j,3).value
        finalValue =insuranceExpiryAmount.get(i) + val
        s2.cell(j,3, value =finalValue)
    else:
        s2.cell(j, 3, value=insuranceExpiryAmount.get(i))
    wb.save("C:\\Users\\vivek\\Desktop\\MR. MAHESH BASUDKAR PROJECTION REPORT1.xlsx")

def insurancAmount(currentYear):
    global insuranceExpiryAmount
    global insuranceExpiryYear
    for i in insuranceExpiryYear.keys():
        expiryYear =insuranceExpiryYear.get(i)
        checkAmount(expiryYear,i,insuranceExpiryAmount,currentYear)

def writeFourthColumnTillRetirement(totalInvestmentInFirstYear,retirementAge,currentAgeInYear):
    listOfValue =[]

    j = 5
    k = 3
    for i in range(currentAgeInYear, 86):
        new_row = j
        new_col = k
        listOfValue.append(s2.cell(new_row, new_col).value)

        j = j + 1

    j = 5
    k = 4
    l=0
    for i in range(currentAgeInYear, 86):
        new_row = j
        new_col = k
        if listOfValue[l] is not None:
            if listOfValue[l] == 0:
                s2.cell(new_row, new_col, value=0)
            else:
                totalInvestmentInFirstYear =totalInvestmentInFirstYear + listOfValue[l]
                s2.cell(new_row, new_col, value=totalInvestmentInFirstYear)
        l = l + 1
        j = j + 1

    wb.save("C:\\Users\\vivek\\Desktop\\MR. MAHESH BASUDKAR PROJECTION REPORT1.xlsx")



#### Value to be calculate for first row differently::::

def writeInvestmentToExcelForFirstYear(totalInvestment):
    j = 4
    k = 3
    for i in range(5):
        new_row = j
        new_col = k
        s2.cell(new_row, new_col, value=totalInvestment)
        k=k+1
        if(i==2):
            k=k+1
    wb.save("C:\\Users\\vivek\\Desktop\\MR. MAHESH BASUDKAR PROJECTION REPORT1.xlsx")
def getFirstYearInvestment(monthlyInvestment, alreadyInvested, freshInvestment,monthOfDob):
    firstInvestmentMonth = 12 - monthOfDob
    if (firstInvestmentMonth == 0):
        monthlyInvestment =monthlyInvestment
    else:
        monthlyInvestment =monthlyInvestment * firstInvestmentMonth
    totalInvestmentInFirstYear =monthlyInvestment + alreadyInvested + freshInvestment
    writeInvestmentToExcelForFirstYear(totalInvestmentInFirstYear)
    return totalInvestmentInFirstYear


#----- Commenting the following methods
"""
needBaseDataDictionary['monthlyInvestment'] = 75000
needBaseDataDictionary['alreadyInvested'] = 1200000
needBaseDataDictionary['freshInvestment'] =2500000

monthlyInvestment =int(needBaseDataDictionary['monthlyInvestment'])
alreadyInvested = int(needBaseDataDictionary['alreadyInvested'])
freshInvestment =int(needBaseDataDictionary['freshInvestment'])
monthOfDob =needBaseDataDictionary['monthOfDob']
totalInvestmentInFirstYear =getFirstYearInvestment(monthlyInvestment, alreadyInvested, freshInvestment,monthOfDob)


retirementAge = needBaseDataDictionary.get("retirementAge")

currentYear = needBaseDataDictionary.get("currentYear")
currentAgeInYear =needBaseDataDictionary.get("currentAgeInYear")
monthlyRequirementAfterRetirement = needBaseDataDictionary.get("monthlyRequirementAfterRetirement")
writeFirstColumnTillRetirement(retirementAge,currentYear,currentAgeInYear)
writeSecondColumnTillRetirement(retirementAge,currentAgeInYear)
writeThirdColumnTillRetirement(monthlyInvestment,retirementAge,currentAgeInYear)
insurancAmount(currentYear)
# to write in 4th column we need totalInvestmentInFirstYear value:
print(totalInvestmentInFirstYear)
writeFourthColumnTillRetirement(totalInvestmentInFirstYear,retirementAge,currentAgeInYear)
"""
#### Yearly withdrawal system.
def writeToYearlyWithdrawal(totalWithdrawalForGraduation,totalWithdrawalForPostGraduation,totalWithdrawalForMarriage,dobOfChild,currentYear):

    childAge  =currentYear - dobOfChild
    gradYear =dobOfChild + 18
    postGradYear =dobOfChild + 22
    marriage = dobOfChild + 25
    j=4
    k=6
    for i in range(dobOfChild+1,dobOfChild+26):
        new_row=j
        new_col=k
        s2.cell(new_row, new_col, value=0)
        if(i==gradYear):
            new_row = 4 + 16
            new_col =k
            existValue=s2.cell(new_row,new_col).value
            if existValue is None:
                existValue =0
            s2.cell(new_row,new_col, value =totalWithdrawalForGraduation+existValue)
        if(i==postGradYear):
            new_row = 4 + 20
            new_col = k
            existValue = s2.cell(new_row, new_col).value
            if existValue is None:
                existValue =0
            s2.cell(new_row, new_col, value=totalWithdrawalForPostGraduation + existValue)
        if(i==marriage):
            new_row = 4 + 23
            new_col = k
            existValue = s2.cell(new_row, new_col).value
            if existValue is None:
                existValue =0
            s2.cell(new_row, new_col, value=totalWithdrawalForMarriage + existValue)
        j=j+1
    wb.save("C:\\Users\\vivek\\Desktop\\MR. MAHESH BASUDKAR PROJECTION REPORT1.xlsx")


#--- Commenting the following code.
"""
numberOfChild = needBaseDataDictionary.get("numberOfChild")
listOfYearlyWithdrawal =[]
if(numberOfChild==1):
    totalWithdrawalForGraduation = needBaseDataDictionary.get("firstChildGraducationCost")
    totalWithdrawalForPostGraduation =needBaseDataDictionary.get("firstChildPostGraduationCost")
    totalWithdrawalForMarriage = needBaseDataDictionary.get("firstChildMarriage")
    dobOfChild =needBaseDataDictionary.get("firstChildDob")
    writeToYearlyWithdrawal(totalWithdrawalForGraduation,totalWithdrawalForPostGraduation,totalWithdrawalForMarriage,dobOfChild,currentYear)
"""
#Montly Retirement Withdrawal
##### Write to Value OF Investment, Balance Amount, Growth Rate####
def compoundCalculation(Principal, Rate, Time):
    Amount =Principal *(pow((1 + Rate/100), Time))
    Amount = int(round(Amount,2))
    return Amount
def writeWithdrawalForRetirement(monthlyRequirementAfterRetirement,retirementAge,currentAgeInYear):
    j=4
    k=6
    Rate =10
    j =j +(retirementAge-currentAgeInYear)
    j=j+1
    count =1
    print("hhhhh4h4h4h4h4h44h4h4h")
    print(j)
    print(retirementAge)
    print(currentAgeInYear)
    for i in range(retirementAge+1,86):
        new_row =j
        new_col =k
        count =count+1
        Amount = compoundCalculation(monthlyRequirementAfterRetirement, Rate, 1)
        existValue = s2.cell(new_row, new_col).value
        #Amount = (Amount + existValue)*12
        if existValue is None:
            existValue = 0
        s2.cell(new_row, new_col, value=(Amount + existValue)*12)
        if count==10:
            Rate=Rate-2
            monthlyRequirementAfterRetirement=Amount
            count =1
        j = j + 1
    wb.save("C:\\Users\\vivek\\Desktop\\MR. MAHESH BASUDKAR PROJECTION REPORT1.xlsx")


# commenting the following code

#writeWithdrawalForRetirement(monthlyRequirementAfterRetirement,retirementAge,currentAgeInYear)


def writeToValueOfInvestment(currentAgeInYear,retirementAge):
    j = 5
    k = 5
    Rate =12
    for i in range(currentAgeInYear+1,86):
        growthRate = s2.cell(j - 1, k + 3).value
        investedAmount = s2.cell(j, k - 2).value
        yearlywithdrawal = s2.cell(j, k + 1).value
        valueOfInvestment = growthRate + investedAmount
        balanceAmount = valueOfInvestment - yearlywithdrawal
        if currentAgeInYear==50:
            Rate =10
        if currentAgeInYear == 60:
            Rate =8
        if currentAgeInYear==70:
            Rate =6
        if currentAgeInYear==80:
            Rate =4
        growthAmount = compoundCalculation(balanceAmount,Rate,1)
        s2.cell(j, k, value=valueOfInvestment)
        s2.cell(j, k+2, value= balanceAmount)
        s2.cell(j, k+3, value =growthAmount)
        j=j+1
    wb.save("C:\\Users\\vivek\\Desktop\\MR. MAHESH BASUDKAR PROJECTION REPORT1.xlsx")
# commenting the following code
#writeToValueOfInvestment(currentAgeInYear,retirementAge)

#inserting a new row
def insertRow(currentAgeInYear,retirementAge):
    j=4
    for i in range(currentAgeInYear,retirementAge):
        j=j+1
    j=j+1
    s2.insert_rows(j,1)
    wb.save("C:\\Users\\vivek\\Desktop\\MR. MAHESH BASUDKAR PROJECTION REPORT1.xlsx")
    for n in range(retirementAge+1,retirementAge+10):
        j=j+1
    j=j+1
    s2.insert_rows(j,1)
    wb.save("C:\\Users\\vivek\\Desktop\\MR. MAHESH BASUDKAR PROJECTION REPORT1.xlsx")

#insertRow(currentAgeInYear,retirementAge)

def methodToFillProjectionReport(rate1,Input,Input2,Input3):
    global needBaseDataDictionary
    global insuranceCostDataDictionary
    global currentYear
    global currentAgeInYear
    print("inside Main mehtod")
    print(getDobOfAllMemebers())
    print(calculateExpenses())
    print(calculateInsuaranceCost())
    print(needBaseDataDictionary)
    # print(insuranceCostDataDictionary)
    print(insuranceExpiryAmountToAdd())
    needBaseDataDictionary['monthlyInvestment'] = 75000
    needBaseDataDictionary['alreadyInvested'] = 1200000
    needBaseDataDictionary['freshInvestment'] = 2500000

    monthlyInvestment = int(needBaseDataDictionary['monthlyInvestment'])
    alreadyInvested = int(needBaseDataDictionary['alreadyInvested'])
    freshInvestment = int(needBaseDataDictionary['freshInvestment'])
    monthOfDob = needBaseDataDictionary['monthOfDob']
    totalInvestmentInFirstYear = getFirstYearInvestment(monthlyInvestment, alreadyInvested, freshInvestment, monthOfDob)

    retirementAge = needBaseDataDictionary.get("retirementAge")

    currentYear = needBaseDataDictionary.get("currentYear")
    currentAgeInYear = needBaseDataDictionary.get("currentAgeInYear")
    monthlyRequirementAfterRetirement = needBaseDataDictionary.get("monthlyRequirementAfterRetirement")
    writeFirstColumnTillRetirement(retirementAge, currentYear, currentAgeInYear)
    writeSecondColumnTillRetirement(retirementAge, currentAgeInYear)
    writeThirdColumnTillRetirement(monthlyInvestment, retirementAge, currentAgeInYear)
    insurancAmount(currentYear)
    # to write in 4th column we need totalInvestmentInFirstYear value:
    print(totalInvestmentInFirstYear)
    writeFourthColumnTillRetirement(totalInvestmentInFirstYear, retirementAge, currentAgeInYear)
    numberOfChild = needBaseDataDictionary.get("numberOfChild")
    listOfYearlyWithdrawal = []
    if (numberOfChild == 1):
        totalWithdrawalForGraduation = needBaseDataDictionary.get("firstChildGraducationCost")
        totalWithdrawalForPostGraduation = needBaseDataDictionary.get("firstChildPostGraduationCost")
        totalWithdrawalForMarriage = needBaseDataDictionary.get("firstChildMarriage")
        dobOfChild = needBaseDataDictionary.get("firstChildDob")
        writeToYearlyWithdrawal(totalWithdrawalForGraduation, totalWithdrawalForPostGraduation,
                                totalWithdrawalForMarriage, dobOfChild, currentYear)
    writeWithdrawalForRetirement(monthlyRequirementAfterRetirement, retirementAge, currentAgeInYear)
    writeToValueOfInvestment(currentAgeInYear, retirementAge)
    print("ending Main mehtod")


methodToFillProjectionReport()